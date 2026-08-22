#!/usr/bin/env python3
"""
Indoor Football Index — team data consolidator.

Builds the shared teams.json that team.html reads from at runtime, keyed
by a slugified team name. Combines two sources from the master workbook:

  - the 'All Teams' tab: bio data (league, years, records, name/location/
    stadium/coach history, championships, logos, info)
  - the 'Schedule' tab: game-by-game results, matched to each team via
    'Current Home Name' / 'Current Away Name' (so a franchise's full
    history follows it across any past rebrands) and pre-grouped into
    seasons here at build time, so team.html never has to fetch or filter
    the full 14,000+ row schedule itself.

Usage:
    python3 build_teams_json.py path/to/2026_IFI.xlsx path/to/teams.json
"""

import sys
import re
import json
import datetime
import openpyxl
from collections import defaultdict


def slug(name):
    return re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")


def esc_none(v):
    return v if v not in (None, "") else None


def norm_league(v):
    """League codes should be uppercase acronyms (IFL, AF2, ...), but the
    source sheets occasionally have stray lowercase entries (e.g. 'af2'
    next to 'AF2'). Normalizing here means every consuming page can match
    on league code without needing its own case-insensitive comparison."""
    return v.strip().upper() if isinstance(v, str) and v.strip() else v


def parse_record_cell(v):
    """Recovers a W-L(-T) value Excel/Sheets silently turned into a date."""
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return f"{v.month}-{v.day}"
    return v


def parse_championships(v):
    if v is None or str(v).strip() == "":
        return 0, []
    if hasattr(v, "year") and hasattr(v, "month"):
        return 1, [str(v.year)]
    parts = [p.strip() for p in re.split(r"[,;]", str(v)) if p.strip()]
    if not parts:
        return 0, []
    count = int(parts[0]) if parts[0].isdigit() else len(parts)
    years = parts[1:] if parts[0].isdigit() else parts
    return count, years


def build_coach_list(coach_history):
    if not coach_history:
        return []
    out = []
    for part in str(coach_history).split(","):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(.*?)\s*\(([^)]+)\)$", part)
        if m:
            out.append({"name": m.group(1).strip(), "tenure": m.group(2).strip()})
        else:
            out.append({"name": part, "tenure": ""})
    return out


def load_all_teams_bio(wb):
    ws = wb["All Teams"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    def get(r, col):
        return r[idx[col]] if col in idx and idx[col] < len(r) else None

    bio = {}
    for r in rows[1:]:
        name = get(r, "Team Name")
        if not name:
            continue
        champ_count, champ_years = parse_championships(get(r, "Championships"))

        logos = []
        current_logo = get(r, "Logo")
        if current_logo:
            logos.append({
                "url": current_logo,
                "caption": esc_none(get(r, "Logo Details")),
                "current": True,
            })
        for n in range(1, 6):
            url = get(r, f"Previous Logos {n}")
            if url:
                logos.append({
                    "url": url,
                    "caption": esc_none(get(r, f"Previous Logos {n} Details")),
                    "current": False,
                })

        bio[name] = {
            "name": name,
            "slug": slug(name),
            "league": norm_league(esc_none(get(r, "League"))),
            "years": esc_none(get(r, "Years")),
            "firstSeason": get(r, "First Season"),
            "regRecord": parse_record_cell(get(r, "Reg W-L-T")),
            "playoffRecord": parse_record_cell(get(r, "Playoff W-L")),
            "totalRecord": parse_record_cell(get(r, "Total W-L-T")),
            "winPct": fmt_pct(get(r, "Win %")),
            "nameHistory": esc_none(get(r, "Name History")),
            "locationHistory": esc_none(get(r, "Location History")),
            "stadiumHistory": esc_none(get(r, "Stadium History")),
            "coaches": build_coach_list(get(r, "Coach History")),
            "championships": champ_count,
            "championshipYears": champ_years,
            "logos": logos,
            "info": esc_none(get(r, "Info")),
            "seasons": {},
        }
    return bio


def load_schedule_by_team(wb, team_names):
    ws = wb["Schedule"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    CUR_HOME_COL = idx.get("Current Home Name", 17)
    CUR_AWAY_COL = CUR_HOME_COL + 1

    games_by_team = defaultdict(list)
    for r in rows[1:]:
        date = r[idx["Date"]]
        if date is None:
            continue
        cur_home = r[CUR_HOME_COL] if CUR_HOME_COL < len(r) else None
        cur_away = r[CUR_AWAY_COL] if CUR_AWAY_COL < len(r) else None
        home_team = r[idx["Home Team"]]
        away_team = r[idx["Away Team"]]
        # If Current Home/Away Name wasn't filled in for this row (common on
        # freshly-added games before that column gets backfilled), fall back
        # to the plain Home/Away Team value rather than silently dropping
        # the game from both teams' pages.
        cur_home = cur_home or home_team
        cur_away = cur_away or away_team
        home_score, home_outcome = parse_score_or_outcome(r[idx["Home Score"]])
        away_score, away_outcome = parse_score_or_outcome(r[idx["Away Score"]])
        if home_score is None and home_outcome is None:
            continue
        if away_score is None and away_outcome is None:
            continue
        week = r[idx["Week"]] if "Week" in idx else None
        week_low = str(week).lower()
        is_championship = "championship" in week_low
        is_playoff = is_championship or any(k in week_low for k in ("playoff", "semifinal", "quarterfinal"))
        year = str(date.year)

        for team_name, is_home in ((cur_home, True), (cur_away, False)):
            if team_name not in team_names:
                continue
            opp = away_team if is_home else home_team
            tie_col = idx.get("Tie")
            tie_flag = (tie_col is not None and r[tie_col] == "Y")

            if home_score is not None and away_score is not None:
                own_score = home_score if is_home else away_score
                opp_score = away_score if is_home else home_score
                wl = "T" if tie_flag or own_score == opp_score else ("W" if own_score > opp_score else "L")
                score_str = f"{fmt_score(own_score)}-{fmt_score(opp_score)}"
            else:
                # outcome-only row (no real score on file) — use the
                # recorded W/L/T text directly rather than guessing
                own_outcome = home_outcome if is_home else away_outcome
                opp_outcome = away_outcome if is_home else home_outcome
                wl = own_outcome or ("T" if tie_flag else ("W" if opp_outcome == "L" else "L"))
                score_str = "\u2014"

            games_by_team[team_name].append({
                "date": date,
                "dateLabel": date.strftime("%b %d"),
                "opp": opp,
                "home": is_home,
                "score": score_str,
                "wl": wl,
                "playoff": is_playoff,
                "championship": is_championship,
                "year": year,
            })
    for t in games_by_team:
        games_by_team[t].sort(key=lambda g: g["date"])
    return games_by_team


def parse_score_or_outcome(v):
    """Most Home/Away Score cells are real numbers. A small number of older
    rows (mostly LFL-era games) only ever recorded the outcome as literal
    'W'/'L'/'T' text instead of a real score. Returns (number_or_None,
    outcome_letter_or_None) so callers can use whichever is available."""
    if v is None:
        return None, None
    if isinstance(v, str):
        s = v.strip().upper()
        if s in ("W", "L", "T"):
            return None, s
        try:
            n = float(s) if "." in s else int(s)
            return n, None
        except ValueError:
            return None, None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return v, None


def fmt_score(v):
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return str(v)


def fmt_pct(v):
    """Excel stores percentages as raw decimals (0.475); the '%' is just a
    display format, which openpyxl doesn't expose via data_only. Convert it
    back to a percent string here so the page doesn't have to guess."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v <= 1:
            v = v * 100
        return f"{v:.1f}%"
    return v


def season_summary(games):
    w = l = t = 0
    for g in games:
        if g["wl"] == "W":
            w += 1
        elif g["wl"] == "L":
            l += 1
        else:
            t += 1
    return f"{w}-{l}" + (f"-{t}" if t else "")


def load_sbs_teams(wb):
    """Reads the 'SBS Teams' tab (Team Name, League, Year, Reg Wins, Reg Loss,
    Ties, Playoff W, Playoff Loss, W-L-T, PO W-L, Conference, Division,
    Result, Current Name, Logo) if present. Matched on Current Name + Year
    so a franchise's season history follows it across any past rebrands,
    same as everywhere else. 'Team Name' and 'Logo' capture that season's
    actual identity (which can differ from the current one after a
    rebrand); both fall back to the current name/logo when left blank.
    Returns {(current_name, year_str): {...}}, or {} if the tab doesn't
    exist yet."""
    if "SBS Teams" not in wb.sheetnames:
        return {}
    ws = wb["SBS Teams"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    def get(r, col):
        return r[idx[col]] if col in idx and idx[col] < len(r) else None

    out = {}
    for r in rows[1:]:
        current_name = get(r, "Current Name")
        year = get(r, "Year")
        if not current_name or year is None:
            continue
        year = str(int(year)) if isinstance(year, (int, float)) else str(year)

        reg = parse_record_cell(get(r, "W-L-T"))
        if not reg:
            w, l, t = get(r, "Reg Wins"), get(r, "Reg Loss"), get(r, "Ties")
            try:
                if w is not None and l is not None:
                    reg = f"{int(w)}-{int(l)}" + (f"-{int(t)}" if t else "")
            except (TypeError, ValueError):
                pass  # corrupted cell (e.g. a stray date/time) — leave reg unset rather than crash

        post = parse_record_cell(get(r, "PO W-L"))
        if not post:
            pw, pl = get(r, "Playoff W"), get(r, "Playoff Loss")
            try:
                if pw is not None and pl is not None and (pw or pl):
                    post = f"{int(pw)}-{int(pl)}"
            except (TypeError, ValueError):
                pass

        logo = esc_none(get(r, "Logo"))
        if logo and str(logo).strip().lower() in ("no match", "n/a", "none", "-"):
            logo = None

        season_name = esc_none(get(r, "Team Name"))

        out[(current_name, year)] = {
            "name": season_name,
            "league": norm_league(esc_none(get(r, "League"))),
            "reg": reg,
            "post": post,
            "conference": esc_none(get(r, "Conference")),
            "division": esc_none(get(r, "Division")),
            "result": esc_none(get(r, "Result")),
            "logo": logo,
        }
    return out


def build_seasons(games):
    by_year = defaultdict(list)
    for g in games:
        by_year[g["year"]].append(g)
    seasons = {}
    for year, yr_games in by_year.items():
        reg_games = [g for g in yr_games if not g["playoff"]]
        post_games = [g for g in yr_games if g["playoff"]]
        seasons[year] = {
            "reg": season_summary(reg_games),
            "post": season_summary(post_games) if post_games else None,
            "games": [
                {
                    "date": g["dateLabel"] + (" \u00b7 CH" if g.get("championship") else (" \u00b7 PO" if g["playoff"] else "")),
                    "opp": ("vs " if g["home"] else "at ") + g["opp"],
                    "score": g["score"].replace("-", "\u2013"),
                    "wl": g["wl"],
                    "playoff": g["playoff"],
                }
                for g in yr_games
            ],
        }
    return seasons


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_teams_json.py <2026_IFI.xlsx> <teams.json>")
        sys.exit(1)
    xlsx_path, out_path = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    bio = load_all_teams_bio(wb)
    print(f"Loaded {len(bio)} teams from 'All Teams' tab")

    games_by_team = load_schedule_by_team(wb, set(bio.keys()))
    matched = 0
    for name, team in bio.items():
        games = games_by_team.get(name, [])
        team["seasons"] = build_seasons(games)
        if games:
            matched += 1
    print(f"Matched schedule data for {matched}/{len(bio)} teams")

    sbs = load_sbs_teams(wb)
    if sbs:
        by_team = defaultdict(list)
        for (cur_name, year), row in sbs.items():
            by_team[cur_name].append((year, row))

        sbs_overrides = 0
        for name, team in bio.items():
            for year, row in by_team.get(name, []):
                sbs_overrides += 1
                existing = team["seasons"].get(year, {"games": []})
                if row["reg"]:
                    existing["reg"] = row["reg"]
                if row["post"]:
                    existing["post"] = row["post"]
                existing["league"] = row["league"]
                existing["conference"] = row["conference"]
                existing["division"] = row["division"]
                existing["result"] = row["result"]
                if row["logo"]:
                    existing["logo"] = row["logo"]
                if row["name"]:
                    existing["name"] = row["name"]
                team["seasons"][year] = existing
        print(f"Applied {sbs_overrides} season records from 'SBS Teams' tab")

    out = list(bio.values())
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, default=str)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
