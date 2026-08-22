#!/usr/bin/env python3
"""
Indoor Football Index — home page generator.

Reads the 'Schedule' tab of a workbook like 2026_IFI.xlsx and builds index.html:
  - Recent games: the 10 most recently completed games before today
  - Upcoming schedule: the 10 next unplayed games from today forward
  - News: the 4 most recent articles from news.json (built separately by
    build_news_json.py from a folder of Word docs) — run that first, or
    this section will just show an empty state.

Usage:
    python3 generate_home_page.py path/to/2026_IFI.xlsx path/to/index.html [path/to/news.json]
"""

import sys
import os
import re
import json
import datetime
import openpyxl


def esc(s):
    if s is None:
        return ""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def fmt_score(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return esc(v)


def load_schedule(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Schedule"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    games = []
    for r in rows[1:]:
        date = r[idx["Date"]]
        if date is None:
            continue
        home = r[idx["Home Team"]]
        away = r[idx["Away Team"]]
        if not home or not away:
            continue
        home_score = r[idx["Home Score"]] if "Home Score" in idx else None
        away_score = r[idx["Away Score"]] if "Away Score" in idx else None
        week = r[idx["Week"]] if "Week" in idx else None
        week_low = str(week).lower()
        is_championship = "championship" in week_low
        is_playoff = is_championship or any(k in week_low for k in ("playoff", "semifinal", "quarterfinal"))
        games.append({
            "date": date,
            "dow": r[idx["Day of the Week"]] if "Day of the Week" in idx else None,
            "home": home,
            "away": away,
            "time": r[idx["Time (CST)"]] if "Time (CST)" in idx else None,
            "week": week,
            "league": r[idx["League"]] if "League" in idx else None,
            "watch": r[idx["Watch"]] if "Watch" in idx else None,
            "home_score": home_score,
            "away_score": away_score,
            "played": home_score is not None and away_score is not None,
            "playoff": is_playoff,
            "championship": is_championship,
        })
    return games


def load_team_logos(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "All Teams" not in wb.sheetnames:
        return {}
    ws = wb["All Teams"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    if "Team Name" not in idx or "Logo" not in idx:
        return {}
    logos = {}
    for r in rows[1:]:
        name = r[idx["Team Name"]]
        logo = r[idx["Logo"]]
        if name and logo:
            logos[name] = logo
    return logos


def team_block(name, score, other_score, played, logos):
    logo_url = logos.get(name)
    logo_html = f'<img src="{esc(logo_url)}" alt="{esc(name)} logo" onerror="this.style.display=\'none\'">' if logo_url else ""
    cls = "team-block"
    if played and score is not None and other_score is not None:
        cls += " winner" if score > other_score else " loser"
    score_html = f'<div class="score">{fmt_score(score)}</div>' if played and score is not None else ""
    return f'<div class="{cls}"><div class="logo">{logo_html}</div><div class="name">{esc(name)}</div>{score_html}</div>'


def game_card(g, logos):
    date_label = g["date"].strftime("%b %-d, %Y" if sys.platform != "win32" else "%b %#d, %Y")
    dow = f' &middot; {esc(g["dow"])}' if g["dow"] else ""
    league = esc(g["league"]) if g["league"] else ""
    card_cls = "game-card playoff" if g["playoff"] else "game-card"
    status_html = '<div class="game-status">Final</div>' if g["played"] else \
        f'<div class="game-status live-time">{esc(g["time"]) if g["time"] else "Time TBD"}</div>'
    watch_html = f'<a class="watch-btn" href="{esc(g["watch"])}" target="_blank" rel="noopener">Watch</a>' if g.get("watch") else "<span></span>"
    week_label = "Championship" if g["championship"] else (esc(g["week"]) if g["week"] else "")
    return f'''    <div class="{card_cls}">
      <div class="game-card-head"><span class="date">{date_label}{dow}</span><span class="league">{league}</span></div>
      {status_html}
      <div class="matchup">{team_block(g["away"], g["away_score"], g["home_score"], g["played"], logos)}<span class="vs-sep">@</span>{team_block(g["home"], g["home_score"], g["away_score"], g["played"], logos)}</div>
      <div class="game-card-foot">{watch_html}<span class="week-label">{week_label}</span></div>
    </div>'''


def load_news(news_json_path):
    if not news_json_path or not os.path.isfile(news_json_path):
        return []
    with open(news_json_path, encoding="utf-8") as f:
        return json.load(f)


def strip_tags(html):
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html or "")).strip()


def build_news_html(articles):
    if not articles:
        return ('  <div class="tba-strip"><div class="big">No news articles yet.</div>'
                '<div class="small">Run build_news_json.py against a folder of Word docs to populate this.</div></div>')
    cards = []
    for item in articles[:4]:
        cls = "news-card feature" if item.get("featured") else "news-card"
        excerpt_len = 260 if item.get("featured") else 160
        excerpt = strip_tags(item.get("body"))
        if len(excerpt) > excerpt_len:
            excerpt = excerpt[:excerpt_len].strip() + "\u2026"
        kicker_html = f'<div class="kicker">{esc(item["kicker"])}</div>' if item.get("kicker") else ""
        cards.append(
            f'    <a class="{cls}" href="article.html?article={esc(item["slug"])}">\n'
            f'      {kicker_html}\n'
            f'      <h3>{esc(item["title"])}</h3>\n'
            f'      <p>{esc(excerpt)}</p>\n'
            f'      <div class="date">{esc(item.get("date", ""))}</div>\n'
            f'    </a>'
        )
    return '  <div class="news-grid">\n' + "\n".join(cards) + "\n  </div>"


TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Indoor Football Index</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Anton&family=IBM+Plex+Mono:wght@400;500;600&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>

  :root{{
    --arena-black:#111316;
    --arena-black-2:#1a1d21;
    --crimson:#D4263A;
    --amber:#E8A33D;
    --turf:#3C6E47;
    --chalk:#F2EFE6;
    --chalk-dim:#c9c6bc;
    --steel:#82868c;
    --steel-line:#2a2d31;
    --card:#1c1f23;
    --card-line:#2e3136;
  }}

  *{{box-sizing:border-box;}}
  html{{scroll-behavior:smooth;}}
  body{{
    margin:0;
    background:var(--arena-black);
    color:var(--chalk);
    font-family:'Inter',system-ui,sans-serif;
    -webkit-font-smoothing:antialiased;
  }}

  .display{{font-family:'Anton',sans-serif; text-transform:uppercase; letter-spacing:0.01em;}}
  .mono{{font-family:'IBM Plex Mono',monospace;}}
  a{{color:inherit; text-decoration:none;}}

  .hashes{{width:100%; height:14px; display:flex; align-items:center;}}
  .hashes svg{{width:100%; height:100%; display:block;}}

  .topbar{{
    width:100%; background:var(--card); border-bottom:1px solid var(--card-line);
    position:sticky; top:0; z-index:100;
  }}
  .topbar-inner{{display:flex; align-items:center; justify-content:space-between; max-width:1400px; margin:0 auto; padding:16px 32px; gap:20px;}}
  .topbar .site{{font-size:13px; letter-spacing:0.14em; text-transform:uppercase; color:var(--steel); white-space:nowrap;}}
  .topbar .site strong{{color:var(--chalk); font-weight:600;}}
  .nav-desktop{{display:flex; gap:22px; font-size:12px; letter-spacing:0.05em; text-transform:uppercase; color:var(--steel); flex-wrap:wrap;}}
  .nav-desktop a:hover{{color:var(--chalk);}}
  .nav-mobile-quick{{display:none; gap:18px; font-size:12px; letter-spacing:0.05em; text-transform:uppercase; color:var(--steel);}}
  .hamburger-btn{{display:none; background:none; border:none; color:var(--chalk); font-size:22px; line-height:1; cursor:pointer; padding:4px 6px;}}
  .mobile-menu{{display:none; flex-direction:column; background:var(--card); border-bottom:1px solid var(--card-line); padding:6px 24px 14px;}}
  .mobile-menu.open{{display:flex;}}
  .mobile-menu a{{padding:11px 0; font-size:13px; color:var(--chalk-dim); border-bottom:1px solid var(--steel-line); text-transform:uppercase; letter-spacing:0.05em;}}
  .mobile-menu a:last-child{{border-bottom:none;}}
  @media (max-width:860px){{
    .nav-desktop{{display:none;}}
    .nav-mobile-quick{{display:flex;}}
    .hamburger-btn{{display:block;}}
  }}

  .hero{{max-width:1160px; margin:0 auto; padding:64px 24px 44px;}}
  .hero .eyebrow{{
    font-size:12px; letter-spacing:0.2em; text-transform:uppercase; color:var(--crimson);
    font-weight:600; margin:0 0 14px;
  }}
  .hero h1{{font-size:clamp(44px,7vw,84px); line-height:0.92; margin:0 0 18px; color:var(--chalk);}}
  .hero p{{font-size:16px; color:var(--steel); max-width:560px; line-height:1.7; margin:0 0 26px;}}
  .hero .cta{{
    display:inline-flex; align-items:center; gap:8px; background:var(--crimson); color:var(--chalk);
    font-size:14px; font-weight:600; padding:13px 22px; border-radius:8px;
  }}
  .hero .cta:hover{{background:#b81f31;}}

  section{{max-width:1160px; margin:0 auto; padding:48px 24px;}}
  .section-head{{display:flex; align-items:baseline; justify-content:space-between; margin-bottom:22px; gap:16px; flex-wrap:wrap;}}
  .section-head h2{{font-size:24px; margin:0; color:var(--chalk);}}
  .section-head .note{{font-size:13px; color:var(--steel);}}

  .game-grid{{display:grid; grid-template-columns:repeat(3,1fr); gap:14px;}}
  @media (max-width:900px){{.game-grid{{grid-template-columns:repeat(2,1fr);}}}}
  @media (max-width:600px){{.game-grid{{grid-template-columns:1fr;}}}}

  .game-card{{background:var(--card); border:1px solid var(--card-line); border-radius:12px; padding:16px;}}
  .game-card.playoff{{border-color:rgba(232,163,61,0.35);}}
  .game-card-head{{display:flex; justify-content:space-between; align-items:baseline; margin-bottom:10px;}}
  .game-card-head .date{{font-size:12px; color:var(--chalk-dim);}}
  .game-card-head .league{{font-size:11px; color:var(--amber); font-family:'IBM Plex Mono',monospace;}}
  .game-status{{text-align:center; font-size:11px; letter-spacing:0.1em; text-transform:uppercase; color:var(--steel); margin-bottom:10px;}}
  .game-status.live-time{{color:var(--turf);}}
  .matchup{{display:flex; align-items:center; justify-content:space-between; gap:10px; margin-bottom:12px;}}
  .team-block{{flex:1; text-align:center; min-width:0;}}
  .team-block .logo{{width:44px; height:44px; margin:0 auto 8px; border-radius:8px; background:var(--arena-black-2); display:flex; align-items:center; justify-content:center; overflow:hidden;}}
  .team-block .logo img{{width:100%; height:100%; object-fit:contain; padding:5px;}}
  .team-block .name{{font-size:12px; color:var(--chalk); font-weight:600; line-height:1.3;}}
  .team-block .score{{font-size:22px; color:var(--amber); font-family:'IBM Plex Mono',monospace; margin-top:6px;}}
  .team-block.loser .score{{color:var(--steel);}}
  .vs-sep{{font-size:11px; color:var(--steel); flex-shrink:0;}}
  .game-card-foot{{display:flex; justify-content:space-between; align-items:center; gap:8px;}}
  .watch-btn{{
    font-size:11px; padding:6px 12px; border-radius:6px; background:var(--arena-black-2); color:var(--chalk-dim);
    border:1px solid var(--card-line); font-weight:600; text-decoration:none;
  }}
  .watch-btn:hover{{color:var(--chalk); border-color:var(--steel);}}
  .week-label{{font-size:11px; color:var(--steel); margin-left:auto;}}

  .tba-strip{{
    border:1px dashed var(--card-line); border-radius:10px; padding:28px 24px; text-align:center;
  }}
  .tba-strip .big{{font-size:15px; color:var(--chalk-dim); margin-bottom:6px;}}
  .tba-strip .small{{font-size:13px; color:var(--steel);}}

  .news-grid{{display:grid; grid-template-columns:1.3fr 1fr 1fr; gap:16px;}}
  .news-card{{
    background:var(--card); border:1px solid var(--card-line); border-radius:10px;
    padding:22px; display:flex; flex-direction:column; gap:10px; transition:border-color .15s;
  }}
  .news-card:hover{{border-color:var(--crimson);}}
  .news-card.feature{{grid-row:span 2; padding:28px;}}
  .news-card .kicker{{font-size:11px; letter-spacing:0.1em; text-transform:uppercase; color:var(--crimson); font-weight:600;}}
  .news-card h3{{font-size:18px; margin:0; color:var(--chalk); line-height:1.35;}}
  .news-card.feature h3{{font-size:26px;}}
  .news-card p{{font-size:14px; color:var(--chalk-dim); line-height:1.65; margin:0; flex:1;}}
  .news-card .date{{font-size:12px; color:var(--steel); font-family:'IBM Plex Mono',monospace; margin-top:auto;}}
  @media (max-width:900px){{
    .news-grid{{grid-template-columns:1fr 1fr;}}
    .news-card.feature{{grid-row:span 1; grid-column:span 2;}}
  }}
  @media (max-width:600px){{
    .news-grid{{grid-template-columns:1fr;}}
    .news-card.feature{{grid-column:span 1;}}
  }}

  footer{{
    max-width:1160px; margin:0 auto; padding:40px 24px 60px; display:flex; justify-content:space-between;
    flex-wrap:wrap; gap:12px; font-size:12px; color:var(--steel);
  }}

  ::selection{{background:var(--crimson); color:var(--chalk);}}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-inner">
    <div class="site"><a href="index.html"><strong>Indoor Football Index</strong></a></div>
    <nav class="nav-desktop">
      <a href="leagues.html">Leagues</a>
      <a href="teams.html">Teams</a>
      <a href="standings.html">Standings</a>
      <a href="players.html">Players</a>
      <a href="awards.html">Awards</a>
      <a href="coaches.html">Coaches</a>
      <a href="schedule.html">Schedule</a>
      <a href="scorigami.html">Scorigami</a>
      <a href="news.html">News</a>
    </nav>
    <nav class="nav-mobile-quick">
      <a href="teams.html">Teams</a>
      <a href="players.html">Players</a>
      <a href="schedule.html">Schedule</a>
    </nav>
    <button class="hamburger-btn" id="hamburgerBtn" aria-label="Open menu">&#9776;</button>
  </div>
  <div class="mobile-menu" id="mobileMenu">
    <a href="leagues.html">Leagues</a>
    <a href="teams.html">Teams</a>
    <a href="standings.html">Standings</a>
    <a href="players.html">Players</a>
    <a href="awards.html">Awards</a>
    <a href="coaches.html">Coaches</a>
    <a href="schedule.html">Schedule</a>
    <a href="scorigami.html">Scorigami</a>
    <a href="news.html">News</a>
  </div>
</div>

<script>
  (function(){{
    var btn = document.getElementById('hamburgerBtn');
    var menu = document.getElementById('mobileMenu');
    if (btn && menu){{
      btn.addEventListener('click', function(){{ menu.classList.toggle('open'); }});
    }}
  }})();
</script>

<header class="hero">
  <p class="eyebrow">Indoor &amp; arena football, archived</p>
  <h1 class="display">Every team.<br>Every season.<br>One index.</h1>
  <p>Franchise histories, record books, and box scores for indoor and arena football &mdash; built from the ground up, one team at a time.</p>
  <a class="cta" href="teams.html">Browse teams &rarr;</a>
</header>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1160 14"><line x1="0" y1="7" x2="1160" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

<section id="schedule">
  <div class="section-head">
    <h2 class="display">Upcoming schedule</h2>
    <p class="note">The next {upcoming_count} games across every league in the archive.</p>
  </div>
{upcoming_html}
</section>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1160 14"><line x1="0" y1="7" x2="1160" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

<section id="recent">
  <div class="section-head">
    <h2 class="display">Recent games</h2>
    <p class="note">The last {recent_count} completed games across every league in the archive.</p>
  </div>
{recent_html}
</section>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1160 14"><line x1="0" y1="7" x2="1160" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

<section id="news">
  <div class="section-head">
    <h2 class="display">News</h2>
    <p class="note"><a href="news.html" style="color:var(--crimson); text-decoration:underline;">See all news &rarr;</a></p>
  </div>
{news_html}
</section>

<footer>
  <span>Indoor Football Index &middot; Team histories, records, and box scores.</span>
  <span>Data compiled from IFL box scores and franchise archives.</span>
</footer>

</body>
</html>
"""


def generate(xlsx_path, out_path, news_json_path=None):
    games = load_schedule(xlsx_path)
    logos = load_team_logos(xlsx_path)
    today = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    completed = [g for g in games if g["played"] and g["date"] <= today]
    completed.sort(key=lambda g: g["date"], reverse=True)
    recent = completed[:10]

    upcoming = [g for g in games if not g["played"] and g["date"] >= today]
    upcoming.sort(key=lambda g: g["date"])
    upcoming = upcoming[:10]

    if recent:
        recent_html = '  <div class="game-grid">\n' + "\n".join(game_card(g, logos) for g in recent) + "\n  </div>"
    else:
        recent_html = ('  <div class="tba-strip"><div class="big">No completed games found before today.'
                        '</div><div class="small">Check that the Schedule tab has scores filled in.</div></div>')

    if upcoming:
        upcoming_html = '  <div class="game-grid">\n' + "\n".join(game_card(g, logos) for g in upcoming) + "\n  </div>"
    else:
        upcoming_html = ('  <div class="tba-strip"><div class="big">No upcoming games found.</div>'
                          '<div class="small">Check that the Schedule tab has future dates without scores.</div></div>')

    articles = load_news(news_json_path or os.path.join(os.path.dirname(os.path.abspath(out_path)), "news.json"))

    html = TEMPLATE.format(
        upcoming_count=len(upcoming),
        recent_count=len(recent),
        upcoming_html=upcoming_html,
        recent_html=recent_html,
        news_html=build_news_html(articles),
    )

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {out_path} ({len(recent)} recent, {len(upcoming)} upcoming, {len(articles)} news articles)")


if __name__ == "__main__":
    if len(sys.argv) not in (3, 4):
        print("Usage: python3 generate_home_page.py <2026_IFI.xlsx> <output.html> [news.json]")
        sys.exit(1)
    news_arg = sys.argv[3] if len(sys.argv) == 4 else None
    generate(sys.argv[1], sys.argv[2], news_arg)
