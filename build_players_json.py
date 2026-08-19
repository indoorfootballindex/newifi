#!/usr/bin/env python3
"""
Indoor Football Index — player data consolidator.

Builds the shared players.json that player.html reads from at runtime,
keyed by a slugified player name. Two input modes, auto-detected:

  - a FOLDER of team workbooks — scans every *.xlsx in it and stacks each
    one's 'Player Stats' tab (reads the team name from that file's own
    Team Info tab)
  - a single MASTER file — one big sheet with every player's stats plus a
    'Team' column (matched by header text, wherever that column sits)

Usage:
    python3 build_players_json.py path/to/team_files_folder path/to/players.json
    python3 build_players_json.py path/to/master.xlsx path/to/players.json
"""

import sys
import os
import re
import glob
import json
import openpyxl


def slug(name):
    return re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")


def coerce_number(v):
    """Excel sometimes stores a numeric-looking cell as text (mixed column
    formatting is common in hand-edited sheets). Convert '15' -> 15,
    '-2' -> -2, '3.5' -> 3.5, so downstream math (which only sums fields
    where typeof === number) actually works. Non-numeric text is left as-is.
    Also normalizes whole-number floats (8.0 -> 8), which Excel produces
    even for genuinely-numeric cells."""
    if isinstance(v, float) and v == int(v):
        return int(v)
    if not isinstance(v, str):
        return v
    s = v.strip()
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return v


def load_team(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    info_ws = wb["Team Info"]
    headers = [c.value for c in info_ws[1]]
    vals = [c.value for c in info_ws[2]]
    info = dict(zip(headers, vals))
    team_name = info.get("Team Name") or os.path.basename(xlsx_path)

    ws = wb["Player Stats"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    stat_headers = rows[0]
    out = []
    for r in rows[1:]:
        row = dict(zip(stat_headers, r))
        name = row.get("Name")
        if not name or str(name).strip() == "":
            continue
        stats = {k: coerce_number(v) for k, v in row.items() if k and k != "Name" and v is not None and v != ""}
        out.append({
            "name": name,
            "slug": slug(name),
            "team": team_name,
            **stats,
        })
    return out


def load_master(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    team_col = None
    for h in idx:
        if str(h).strip().lower() == "team":
            team_col = h
            break
    if team_col is None:
        print("ERROR: no 'Team' column found in the master file's header row.")
        print(f"Headers found: {list(idx.keys())}")
        sys.exit(1)

    out = []
    for r in rows[1:]:
        row = dict(zip(headers, r))
        name = row.get("Name")
        if not name or str(name).strip() == "":
            continue
        team_name = row.get(team_col)
        if not team_name:
            continue
        stats = {
            k: coerce_number(v) for k, v in row.items()
            if k and k not in ("Name", team_col) and v is not None and v != ""
        }
        out.append({
            "name": name,
            "slug": slug(name),
            "team": team_name,
            **stats,
        })
    return out


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_players_json.py <folder_of_team_xlsx OR master.xlsx> <players.json>")
        sys.exit(1)
    src, out_path = sys.argv[1], sys.argv[2]

    if os.path.isdir(src):
        xlsx_files = sorted(glob.glob(os.path.join(src, "*.xlsx")))
        if not xlsx_files:
            print(f"No .xlsx files found in {src}")
            sys.exit(1)
        all_rows = []
        for path in xlsx_files:
            try:
                rows = load_team(path)
            except KeyError as e:
                print(f"  skipped {os.path.basename(path)} (missing tab: {e})")
                continue
            all_rows.extend(rows)
            print(f"  {os.path.basename(path)}: {len(rows)} player-seasons")
    elif os.path.isfile(src):
        all_rows = load_master(src)
        print(f"  {os.path.basename(src)}: {len(all_rows)} player-seasons across {len({r['team'] for r in all_rows})} teams")
    else:
        print(f"Not found: {src}")
        sys.exit(1)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False)

    unique_players = len({r["slug"] for r in all_rows})
    print(f"Wrote {out_path}: {len(all_rows)} player-season rows, {unique_players} unique players")


if __name__ == "__main__":
    main()
