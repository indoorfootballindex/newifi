#!/usr/bin/env python3
"""
Indoor Football Index — league data builder.

Builds leagues.json for leagues.html (directory) and league.html (detail
page), from Leagues.xlsx:
  - 'Leagues' tab: bio (name, acronym, active, years, logo, commissioner,
    history)
  - 'Season Info' tab: per-league-per-year awards and results (matched on
    Acronym), covering Games Played, Championship, MVP, and a long list of
    other year-end awards, plus a free-text 'About' blurb for that season

Usage:
    python3 build_leagues_json.py path/to/Leagues.xlsx path/to/leagues.json
"""

import sys
import re
import json
import openpyxl
from collections import defaultdict

AWARD_COLUMNS = [
    "MVP", "Ironman of the Year", "Coach of the Year", "Offensive Player of the Year",
    "Defensive Player of the Year", "Special Teams Player of the Year",
    "Offensive Rookie of the Year", "Defensive Rookie of the Year",
    "Assistant Coach of the Year", "Lineman of the Year", "Executive of the Year",
    "President's Award", "Commissioner's Award", "Adam Pringle Award", "Most Improved Player",
]


def slug(name):
    return re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")


def esc_none(v):
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def load_leagues(wb):
    ws = wb["Leagues"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    def get(r, col):
        return r[idx[col]] if col in idx and idx[col] < len(r) else None

    out = {}
    for r in rows[1:]:
        name = get(r, "League Name")
        acronym = get(r, "Acronym")
        if not name or not acronym:
            continue
        out[acronym] = {
            "name": name,
            "slug": slug(name),
            "acronym": acronym,
            "active": str(get(r, "Active?")).strip().upper() == "Y",
            "years": esc_none(get(r, "Years Active")),
            "logo": esc_none(get(r, "Logo")),
            "commissioner": esc_none(get(r, "Commish")),
            "history": esc_none(get(r, "History")),
            "seasons": {},
        }
    return out


def load_season_info(wb):
    ws = wb["Season Info"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    def get(r, col):
        return r[idx[col]] if col in idx and idx[col] < len(r) else None

    by_acronym = defaultdict(dict)
    for r in rows[1:]:
        acronym = get(r, "Acronym")
        year = get(r, "Year")
        if not acronym or year is None:
            continue
        year = str(int(year)) if isinstance(year, (int, float)) else str(year)

        awards = {}
        for col in AWARD_COLUMNS:
            v = esc_none(get(r, col))
            if v:
                awards[col] = v

        by_acronym[acronym][year] = {
            "gamesPlayed": esc_none(get(r, "Games Played")),
            "championship": esc_none(get(r, "Championship")),
            "about": esc_none(get(r, "About")),
            "awards": awards,
        }
    return by_acronym


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_leagues_json.py <Leagues.xlsx> <leagues.json>")
        sys.exit(1)
    xlsx_path, out_path = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    leagues = load_leagues(wb)
    season_info = load_season_info(wb)

    matched = 0
    for acronym, seasons in season_info.items():
        if acronym in leagues:
            leagues[acronym]["seasons"] = seasons
            matched += 1

    unmatched_seasons = [a for a in season_info if a not in leagues]

    out = list(leagues.values())
    out.sort(key=lambda lg: lg["name"])

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, default=str)

    print(f"Wrote {out_path}: {len(out)} leagues, {matched} with season/award data")
    if unmatched_seasons:
        print(f"  Note: Season Info has {len(unmatched_seasons)} acronym(s) with no matching league row: {unmatched_seasons}")


if __name__ == "__main__":
    main()
