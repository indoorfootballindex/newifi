#!/usr/bin/env python3
"""
Indoor Football Index — coaches data builder.

Reads Coaches.xlsx ('Coaches' tab: Name, Active, Reg Wins, Reg Loss, Post
Win, Post Loss, Total, Team History, Picture, Bio; 'Seasons' tab: Name,
Year, Team, Reg Wins, Reg Loss, Post Win, Post Loss) and builds
coaches.json.

Coaches don't get their own separate profile page — player.html reads
this file too and merges it onto the same profile as any matching player
(by name), so someone who both played and later coached gets one unified
page instead of two. Team names in this sheet are often short-form
("Bay Area" instead of "Bay Area Panthers"), so they're resolved against
teams.json the same way Awards.xlsx's short names are.

Usage:
    python3 build_coaches_json.py path/to/Coaches.xlsx path/to/coaches.json [path/to/teams.json]
"""

import sys
import os
import re
import json
import openpyxl


def slug(name):
    return re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")


def fmt_num(v):
    if v is None:
        return 0
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return v


def build_team_resolver(teams_json_path):
    """Same short-name-to-real-team resolution used for Awards.xlsx."""
    if not teams_json_path or not os.path.isfile(teams_json_path):
        return lambda name: (name, slug(name) if name else None)

    with open(teams_json_path, encoding="utf-8") as f:
        teams = json.load(f)

    exact = {t["name"].strip().lower(): t for t in teams}
    by_length = sorted(teams, key=lambda t: -len(t["name"]))

    def resolve(name):
        if not name:
            return None, None
        key = str(name).strip().lower()
        if key in exact:
            return exact[key]["name"], exact[key]["slug"]
        for t in by_length:
            if t["name"].strip().lower().startswith(key):
                return t["name"], t["slug"]
        return name, slug(name)  # no match — fall back to what's on file

    return resolve


def load_seasons(wb, resolve_team):
    if "Seasons" not in wb.sheetnames:
        return {}
    ws = wb["Seasons"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    def get(r, col):
        return r[idx[col]] if col in idx and idx[col] < len(r) else None

    by_name = {}
    for r in rows[1:]:
        name = get(r, "Name")
        year = get(r, "Year")
        if not name or year is None:
            continue
        year = str(int(year)) if isinstance(year, (int, float)) else str(year)
        team_name, team_slug = resolve_team(get(r, "Team"))
        reg_w, reg_l = fmt_num(get(r, "Reg Wins")), fmt_num(get(r, "Reg Loss"))
        post_w, post_l = fmt_num(get(r, "Post Win")), fmt_num(get(r, "Post Loss"))
        by_name.setdefault(name, {})[year] = {
            "team": team_name,
            "teamSlug": team_slug,
            "reg": f"{reg_w}-{reg_l}",
            "post": f"{post_w}-{post_l}" if (post_w or post_l) else None,
        }
    return by_name


def main():
    if len(sys.argv) not in (3, 4):
        print("Usage: python3 build_coaches_json.py <Coaches.xlsx> <coaches.json> [teams.json]")
        sys.exit(1)
    xlsx_path, out_path = sys.argv[1], sys.argv[2]
    teams_json_path = sys.argv[3] if len(sys.argv) == 4 else None
    resolve_team = build_team_resolver(teams_json_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Coaches"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}

    def get(r, col):
        return r[idx[col]] if col in idx and idx[col] < len(r) else None

    seasons_by_name = load_seasons(wb, resolve_team)

    coaches = []
    for r in rows[1:]:
        name = get(r, "Name")
        if not name:
            continue
        active_val = get(r, "Active")
        is_active = bool(active_val) and str(active_val).strip().lower() != "no"
        active_team_name, active_team_slug = (resolve_team(active_val) if is_active else (None, None))

        reg_w, reg_l = fmt_num(get(r, "Reg Wins")), fmt_num(get(r, "Reg Loss"))
        post_w, post_l = fmt_num(get(r, "Post Win")), fmt_num(get(r, "Post Loss"))

        team_history_raw = get(r, "Team History") or ""
        team_history = []
        for part in team_history_raw.split(","):
            part = part.strip()
            if not part:
                continue
            t_name, t_slug = resolve_team(part)
            team_history.append({"name": t_name, "slug": t_slug})

        coaches.append({
            "name": name,
            "slug": slug(name),
            "active": is_active,
            "activeTeam": active_team_name,
            "activeTeamSlug": active_team_slug,
            "regRecord": f"{reg_w}-{reg_l}",
            "postRecord": f"{post_w}-{post_l}" if (post_w or post_l) else None,
            "totalRecord": get(r, "Total") or f"{reg_w+post_w}-{reg_l+post_l}",
            "teamHistory": team_history,
            "picture": get(r, "Picture"),
            "bio": get(r, "Bio"),
            "seasons": seasons_by_name.get(name, {}),
        })

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(coaches, f, ensure_ascii=False)

    active_count = sum(1 for c in coaches if c["active"])
    print(f"Wrote {out_path}: {len(coaches)} coaches ({active_count} currently active)")


if __name__ == "__main__":
    main()
