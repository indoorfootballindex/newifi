#!/usr/bin/env python3
"""
Indoor Football Index — team page generator.

Reads a team workbook ('Team Info', 'Player Stats', 'Roster', 'Franchise
Record Holders' tabs) plus the master schedule workbook (its 'Schedule' tab,
e.g. 2026_IFI.xlsx) and writes a self-contained HTML team page in the house
style. Game history now comes from the shared master schedule rather than a
per-team 'Game History' tab, so it only has to be entered once.

Games are attributed to a team using the 'Current Home Name' /
'Current Away Name' columns in the Schedule tab, so a franchise's full
history follows it across any past rebrands. The opponent shown in each
game row still uses whatever name that opponent played under at the time.

Usage:
    python3 generate_team_page.py path/to/TeamSheet.xlsx path/to/2026_IFI.xlsx path/to/output.html
"""

import sys
import re
import json
import openpyxl
from collections import defaultdict

# ---------------------------------------------------------------------------
# Column -> record label mapping (Player Stats tab, current-season leaders)
# ---------------------------------------------------------------------------
LEADER_CATEGORIES = [
    ("offense", "Passing yards", "P Yds", "yds"),
    ("offense", "Passing touchdowns", "P TD", "td"),
    ("offense", "Rushing yards", "Rush Yds", "yds"),
    ("offense", "Rushing touchdowns", "Rush TD", "td"),
    ("offense", "Receiving yards", "Rec Yds", "yds"),
    ("offense", "Receiving touchdowns", "Rec TD", "td"),
    ("defense", "Tackles", "TotTkl", "tkl"),
    ("defense", "Sacks", "Sck", "sck"),
    ("defense", "Interceptions", "Def Int", "int"),
    ("defense", "Passes defended", "PBU", "pbu"),
    ("defense", "Tackles for loss", "TFL", "tfl"),
    ("defense", "Forced fumbles", "FF", "ff"),
    ("special", "Kickoff return yards", "KR Yds", "yds"),
    ("special", "Field goals made", "FGM", "fgm"),
    ("special", "PATs made", "XPM", "xpm"),
]

# Franchise Record Holders tab: keyword -> unit bucket
RECORD_UNIT_RULES = [
    ("defense", ["tackle", "sack", "interception", "fumble", "passes defended", "broken up"]),
    ("special", ["field goal", "pat", "kickoff", "blocked", "punt"]),
]
def classify_record(label):
    low = label.lower()
    for unit, keywords in RECORD_UNIT_RULES:
        if any(k in low for k in keywords):
            return unit
    return "offense"

# Sub-grouping within each unit, for the grouped table headers
RECORD_GROUPS = {
    "passing": "Passing", "rushing": "Rushing", "receiving": "Receiving", "reception": "Receiving",
    "all-purpose": "All-purpose", "total td": "All-purpose",
    "points": "All-purpose", "games played": "All-purpose",
    "tackle": "Tackling", "tackles for loss": "Tackling",
    "sack": "Pass rush", "interception": "Coverage", "passes defended": "Coverage",
    "forced fumble": "Takeaways", "fumble recover": "Takeaways",
    "field goal": "Kicking", "pat": "Kicking", "blocked": "Kicking",
    "kickoff": "Returns", "punt": "Returns",
}
def record_group(label):
    low = label.lower()
    for k, v in RECORD_GROUPS.items():
        if k in low:
            return v
    return "Other"


def esc(s):
    if s is None:
        return ""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# Player pages are served by ONE template (player.html) that reads a
# ?player=slug query param at runtime and looks the player up in
# players.json — not one static file per player. Edit the base path below
# once player.html is deployed at its real location.
PLAYER_PAGE_BASE = "player.html?player="


def player_slug(name):
    slug = re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")
    return slug


def player_link(name):
    if not name:
        return ""
    url = PLAYER_PAGE_BASE + player_slug(name)
    return f'<a href="{esc(url)}" class="plain-link">{esc(name)}</a>'


def parse_championships(v):
    """Accepts '#, Year, Year...' or '#;Year;Year...' (e.g. '2, 2021, 2024'),
    a plain count, or a datetime (when Sheets/Excel silently auto-converted a
    value like '1;2023' into a date). Returns (count_str, [years])."""
    if v is None or str(v).strip() == "":
        return "0", []
    if hasattr(v, "year") and hasattr(v, "month"):
        # Got silently turned into a date by the spreadsheet. The count is
        # unrecoverable at that point, so assume one championship in that year.
        return "1", [str(v.year)]
    parts = [p.strip() for p in re.split(r"[,;]", str(v)) if p.strip()]
    if not parts:
        return "0", []
    count = parts[0] if parts[0].isdigit() else str(len(parts) - 1) if len(parts) > 1 else "0"
    years = [p for p in parts[1:] if p]
    return count, years


def js_str(s):
    return json.dumps(s if s is not None else "")


def fmt_num(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        v = int(v)
    if isinstance(v, int):
        return f"{v:,}"
    return f"{v:,}" if isinstance(v, (int, float)) else str(v)


def parse_record_cell(v):
    """Handles the 'W-L' text case and the case where Sheets/Excel silently
    turned '3-3' into a date. Returns a plain 'W-L' string or None."""
    if v is None:
        return None
    if hasattr(v, "month") and hasattr(v, "day"):
        return f"{v.month}-{v.day}"
    return str(v)


def load_team_info(wb):
    ws = wb["Team Info"]
    headers = [c.value for c in ws[1]]
    vals = [c.value for c in ws[2]]
    info = dict(zip(headers, vals))
    info["Reg. Record"] = parse_record_cell(info.get("Reg. Record"))
    info["Post. Record"] = parse_record_cell(info.get("Post. Record"))
    info["Record"] = parse_record_cell(info.get("Record"))
    return info


def load_game_history(schedule_wb, team_name):
    """Reads the master Schedule tab and returns every game for a franchise's
    CURRENT identity, regardless of what name the team played under at the
    time. Matching is done on 'Current Home Name' (column R) and the
    unlabeled 'Current Away Name' column immediately after it (column S) —
    that second header is blank in the source sheet, so it's read by fixed
    position rather than by name."""
    ws = schedule_wb["Schedule"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    CUR_HOME_COL = idx.get("Current Home Name", 17)
    CUR_AWAY_COL = CUR_HOME_COL + 1  # blank header, immediately follows Current Home Name

    seasons = defaultdict(list)
    for r in rows[1:]:
        date = r[idx["Date"]]
        if date is None:
            continue
        cur_home = r[CUR_HOME_COL] if CUR_HOME_COL < len(r) else None
        cur_away = r[CUR_AWAY_COL] if CUR_AWAY_COL < len(r) else None
        if cur_home != team_name and cur_away != team_name:
            continue
        home_team = r[idx["Home Team"]]
        away_team = r[idx["Away Team"]]
        home_score = r[idx["Home Score"]]
        away_score = r[idx["Away Score"]]
        if home_score is None or away_score is None:
            continue  # unplayed game — not part of the historical record yet
        is_home = (cur_home == team_name)
        opp = away_team if is_home else home_team  # historical opponent name at the time
        own_score = home_score if is_home else away_score
        opp_score = away_score if is_home else home_score
        tie = r[idx.get("Tie")] == "Y" if "Tie" in idx else False
        wl = "T" if tie else ("W" if own_score > opp_score else "L")
        score_str = f"{fmt_num(own_score)}-{fmt_num(opp_score)}"
        week = r[idx["Week"]] if "Week" in idx else None
        week_low = str(week).lower()
        is_playoff = any(k in week_low for k in ("playoff", "championship", "semifinal", "quarterfinal"))
        year = str(date.year)
        seasons[year].append({
            "date": date, "date_label": date.strftime("%b %d"),
            "opp": opp, "home": is_home, "score": score_str,
            "wl": wl, "playoff": is_playoff,
        })
    for y in seasons:
        seasons[y].sort(key=lambda g: g["date"])
    return seasons


def load_season_table(wb, team_name):
    """Reads the 'Season' tab (Team Name, League, Year, Record, Playoff
    Record, Conference, Division, Result, Current Name, Logo) from the
    team's own workbook, if present. Matches rows on 'Current Name' so a
    franchise's full history follows it across any past rebrands — same
    pattern as the master Schedule tab. Returns {year: {...}} keyed by
    string year, or None if the tab doesn't exist yet — callers should fall
    back to computing records from the master Schedule tab in that case."""
    if "Season" not in wb.sheetnames:
        return None
    ws = wb["Season"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    if "Year" not in idx or "Record" not in idx:
        return None
    match_col = "Current Name" if "Current Name" in idx else "Team Name"
    if match_col not in idx:
        return None
    out = {}
    for r in rows[1:]:
        if r[idx[match_col]] != team_name:
            continue
        year = r[idx["Year"]]
        if year is None:
            continue
        year = str(int(year)) if isinstance(year, float) else str(year)
        out[year] = {
            "league": r[idx["League"]] if "League" in idx else None,
            "record": parse_record_cell(r[idx["Record"]]),
            "playoff_record": parse_record_cell(r[idx["Playoff Record"]]) if "Playoff Record" in idx else None,
            "conference": r[idx["Conference"]] if "Conference" in idx else None,
            "division": r[idx["Division"]] if "Division" in idx else None,
            "result": r[idx["Result"]] if "Result" in idx else None,
        }
    return out


def season_summary(games):
    reg_w = reg_l = reg_t = post_w = post_l = post_t = 0
    for g in games:
        if g["playoff"]:
            post_w += g["wl"] == "W"
            post_l += g["wl"] == "L"
            post_t += g["wl"] == "T"
        else:
            reg_w += g["wl"] == "W"
            reg_l += g["wl"] == "L"
            reg_t += g["wl"] == "T"
    reg = f"{reg_w}-{reg_l}" + (f"-{reg_t}" if reg_t else "")
    post = (f"{post_w}-{post_l}" + (f"-{post_t}" if post_t else "")) if (post_w + post_l + post_t) else None
    return reg, post


def first_playoff_win_year(seasons):
    for y in sorted(seasons.keys()):
        for g in seasons[y]:
            if g["playoff"] and g["wl"] == "W":
                return y
    return None


def current_streak(seasons):
    """Most recent completed game backwards, across all seasons combined."""
    all_games = []
    for y in seasons:
        all_games.extend(seasons[y])
    all_games.sort(key=lambda g: g["date"])
    if not all_games:
        return None
    streak_type = all_games[-1]["wl"]
    count = 0
    for g in reversed(all_games):
        if g["wl"] == streak_type:
            count += 1
        else:
            break
    return f"{streak_type}{count}"


def load_records(wb):
    ws = wb["Franchise Record Holders"]
    rows = list(ws.iter_rows(values_only=True))[1:]
    grouped = defaultdict(lambda: defaultdict(list))
    for label, name, stat in rows:
        if not label:
            continue
        unit = classify_record(label)
        group = record_group(label)
        clean_label = re.sub(r"^Most\s+", "", label)
        grouped[unit][group].append((clean_label, name, stat))
    return grouped


def load_current_leaders(wb):
    ws = wb["Player Stats"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    data_rows = rows[1:]
    seasons_present = sorted({r[idx["Season"]] for r in data_rows if r[idx["Season"]] is not None})
    if not seasons_present:
        return None, [], None
    latest = seasons_present[-1]
    latest_display = int(latest) if isinstance(latest, float) and latest == int(latest) else latest
    latest_rows = [r for r in data_rows if r[idx["Season"]] == latest]
    leaders = []
    for unit, label, col, suffix in LEADER_CATEGORIES:
        if col not in idx:
            continue
        best_name, best_val = None, None
        for r in latest_rows:
            v = r[idx[col]]
            if v is None:
                continue
            if best_val is None or v > best_val:
                best_val, best_name = v, r[idx["Name"]]
        if best_name is not None:
            leaders.append((unit, label, best_name, fmt_num(best_val), suffix))
    return latest_display, leaders, latest


def build_records_table(grouped):
    unit_order = ["offense", "defense", "special"]
    out = []
    for i, unit in enumerate(unit_order):
        groups = grouped.get(unit, {})
        style = "" if i == 0 else ' style="display:none"'
        out.append(f'    <tbody data-unit="{unit}"{style}>')
        for group_name, records in groups.items():
            out.append(f'      <tr class="group-row"><td colspan="3">{esc(group_name)}</td></tr>')
            for label, name, stat in records:
                out.append(
                    f'      <tr><td class="rec">{esc(label)}</td>'
                    f'<td class="holder">{player_link(name)}</td>'
                    f'<td class="val">{esc(fmt_num(stat))}</td></tr>'
                )
        out.append('    </tbody>')
    return "\n".join(out)


def build_leader_cards(leaders):
    unit_order = ["offense", "defense", "special"]
    out = []
    for i, unit in enumerate(unit_order):
        for u, label, name, val, suffix in leaders:
            if u != unit:
                continue
            display = "" if i == 0 else ' style="display:none"'
            out.append(
                f'    <div class="leader-card" data-unit="{unit}"{display}>'
                f'<div class="cat">{esc(label)}</div>'
                f'<div class="who">{player_link(name)}</div>'
                f'<div class="stat">{esc(val)}<span>{esc(suffix)}</span></div>'
                f'</div>'
            )
    return "\n".join(out)


def resolve_season(y, seasons, season_table):
    """Prefer the hand-verified Season by Season table when a row exists for
    this year; fall back to computing from game rows otherwise."""
    table_row = season_table.get(y) if season_table else None
    if table_row and table_row.get("record"):
        reg = str(table_row["record"])
        post_raw = table_row.get("playoff_record")
        post = str(post_raw) if post_raw else None
        return reg, post, table_row
    reg, post = season_summary(seasons.get(y, []))
    return reg, post, None


def build_season_js(seasons, season_table=None):
    all_years = set(seasons.keys())
    if season_table:
        all_years |= set(season_table.keys())
    years = sorted(all_years)
    parts = []
    for y in years:
        games = seasons.get(y, [])
        reg, post, table_row = resolve_season(y, seasons, season_table)
        if table_row:
            result_text = table_row.get("result")
            note = str(result_text) if result_text else (
                f"Regular season {reg}." + (f" Postseason {post}." if post else "")
            )
            league = table_row.get("league")
            conference = table_row.get("conference")
            prefix_bits = [b for b in (league, conference) if b]
            if prefix_bits:
                note = " \u00b7 ".join(str(b) for b in prefix_bits) + ". " + note
        else:
            note = f"Regular season {reg}." + (f" Postseason {post}." if post else "")
        entries = []
        for g in games:
            date_label = g["date_label"] + (" \u00b7 PO" if g["playoff"] else "")
            opp_label = ("vs " if g["home"] else "at ") + g["opp"]
            score = g["score"].replace("-", "\u2013")
            entries.append(json.dumps([date_label, opp_label, score, g["wl"]]))
        games_js = ",\n      ".join(entries)
        post_js = f"Postseason: {post.replace('-', chr(8211))}" if post else ""
        parts.append(
            f'    {js_str(y)}: {{rec:{js_str(reg.replace("-", chr(8211)))}, '
            f'note:{js_str(note)}, post:{js_str(post_js)}, games:[\n      {games_js}\n    ]}}'
        )
    return ",\n".join(parts)


def build_year_options(seasons, season_table=None):
    all_years = set(seasons.keys())
    if season_table:
        all_years |= set(season_table.keys())
    years = sorted(all_years, reverse=True)
    out = []
    for i, y in enumerate(years):
        sel = " selected" if i == 0 else ""
        out.append(f'      <option value="{y}"{sel}>{y}</option>')
    return "\n".join(out), years[0] if years else ""


def load_logo_history(info):
    """Scans Team Info for 'Current Logo'/'Current Logo Info'/'Current Logo
    Name' plus any number of 'Previous Logo'/'Previous Logo N' columns
    (tolerant of the 'Prevoius' typo and 'History' vs 'Info' naming seen in
    real sheets), paired up by matching numeric suffix. A logo's name label
    falls back to the team's current name if no override is given — most
    logos won't have had a different team name attached. Returns an ordered
    list of {url, name, caption, current} dicts, current logo first."""
    entries = []
    team_name = info.get("Team Name") or ""
    current_url = info.get("Current Logo")
    if current_url:
        entries.append({
            "url": current_url,
            "name": info.get("Current Logo Name") or team_name,
            "caption": info.get("Current Logo Info") or "",
            "current": True,
        })

    logo_re = re.compile(r"^previous\s+logo\s*(\d*)$", re.IGNORECASE)
    info_re = re.compile(r"^prev(?:ious|oius)\s+logo\s+(?:info|history)\s*(\d*)$", re.IGNORECASE)
    name_re = re.compile(r"^prev(?:ious|oius)\s+logo\s+name\s*(\d*)$", re.IGNORECASE)

    logo_cols = {}
    info_cols = {}
    name_cols = {}
    for key in info:
        if not key:
            continue
        m = logo_re.match(key)
        if m:
            logo_cols[m.group(1)] = key
            continue
        m = info_re.match(key)
        if m:
            info_cols[m.group(1)] = key
            continue
        m = name_re.match(key)
        if m:
            name_cols[m.group(1)] = key

    def sort_key(suffix):
        return (0, 0) if suffix == "" else (1, int(suffix))

    previous = []
    for suffix in sorted(logo_cols.keys(), key=sort_key):
        url = info.get(logo_cols[suffix])
        if not url:
            continue
        caption = info.get(info_cols[suffix]) if suffix in info_cols else ""
        name = info.get(name_cols[suffix]) if suffix in name_cols else None
        previous.append({"url": url, "name": name or team_name, "caption": caption or "", "current": False})

    entries.extend(previous)
    return entries


def build_logo_history(entries):
    if not entries:
        return ""
    out = []
    for e in entries:
        badge = f'<span class="badge">Current</span>' if e["current"] else ""
        out.append(
            f'    <div class="logo-card">'
            f'<div class="badge-slot">{badge}</div>'
            f'<div class="thumb"><img src="{esc(e["url"])}" alt="{esc(e["name"])} logo"></div>'
            f'<div class="logo-name">{esc(e["name"])}</div>'
            f'<div class="caption">{esc(e["caption"])}</div>'
            f'</div>'
        )
    return "\n".join(out)


def build_coaches(coach_history):
    if not coach_history:
        return ""
    out = []
    for part in coach_history.split(","):
        part = part.strip()
        m = re.match(r"^(.*?)\s*\(([^)]+)\)$", part)
        if m:
            name, tenure = m.group(1).strip(), m.group(2).strip()
        else:
            name, tenure = part, ""
        out.append(
            f'    <div class="coach-card"><div class="name">{esc(name)}</div>'
            f'<div class="tenure mono">{esc(tenure)}</div></div>'
        )
    return "\n".join(out)


TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{team_name} — Indoor Football Index</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Anton&family=IBM+Plex+Mono:wght@400;500;600&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>

  :root{{
    --arena-black:#111316;
    --arena-black-2:#1a1d21;
    --crimson:#D4263A;
    --crimson-dim:#8f1c29;
    --amber:#E8A33D;
    --turf:#3C6E47;
    --turf-dim:#2a4d32;
    --chalk:#F2EFE6;
    --chalk-dim:#c9c6bc;
    --steel:#82868c;
    --steel-line:#2a2d31;
    --card:#1c1f23;
    --card-line:#2e3136;
  }}

  *{{box-sizing:border-box;}}
  html{{scroll-behavior:smooth;}}
  body{{
    margin:0;
    background:var(--arena-black);
    color:var(--chalk);
    font-family:'Inter',system-ui,sans-serif;
    -webkit-font-smoothing:antialiased;
  }}

  .display{{font-family:'Anton',sans-serif; text-transform:uppercase; letter-spacing:0.01em;}}
  .mono{{font-family:'IBM Plex Mono',monospace;}}

  a{{color:inherit;}}
  .plain-link{{color:inherit; text-decoration:none;}}
  .plain-link:hover{{text-decoration:none;}}

  .hashes{{width:100%; height:14px; display:flex; align-items:center;}}
  .hashes svg{{width:100%; height:100%; display:block;}}

  .topbar{{display:flex; align-items:center; justify-content:space-between; max-width:1100px; margin:0 auto; padding:22px 24px 0;}}
  .topbar .site{{font-size:13px; letter-spacing:0.14em; text-transform:uppercase; color:var(--steel);}}
  .topbar .site strong{{color:var(--chalk); font-weight:600;}}
  .topbar nav{{display:flex; gap:28px; font-size:13px; letter-spacing:0.06em; text-transform:uppercase; color:var(--steel);}}
  .topbar nav a{{text-decoration:none;}}
  .topbar nav a:hover{{color:var(--chalk);}}

  .hero{{max-width:1100px; margin:0 auto; padding:56px 24px 40px; display:grid; grid-template-columns:auto 1fr auto; gap:36px; align-items:center;}}
  .hero-logo{{width:132px; height:132px; border-radius:50%; border:1px solid var(--card-line); background:var(--card); display:flex; align-items:center; justify-content:center; overflow:hidden; flex-shrink:0;}}
  .hero-logo img{{width:100%; height:100%; object-fit:contain; padding:14px;}}
  .hero-id .eyebrow{{font-size:12px; letter-spacing:0.18em; text-transform:uppercase; color:var(--crimson); font-weight:600; margin:0 0 10px;}}
  .hero-id h1{{font-size:clamp(40px,6vw,72px); line-height:0.92; margin:0 0 14px; color:var(--chalk);}}
  .hero-id .sub{{display:flex; flex-wrap:wrap; gap:6px 18px; font-size:14px; color:var(--steel);}}
  .hero-id .sub span{{white-space:nowrap;}}
  .hero-id .sub b{{color:var(--chalk-dim); font-weight:500;}}
  .hero-record{{text-align:right; border-left:1px solid var(--card-line); padding-left:36px;}}
  .hero-record .label{{font-size:11px; letter-spacing:0.16em; text-transform:uppercase; color:var(--steel); margin:0 0 6px;}}
  .hero-record .num{{font-size:56px; color:var(--chalk); line-height:1;}}
  .hero-record .num span{{color:var(--steel-line);}}
  .hero-record .breakdown{{margin-top:10px; font-size:12px; color:var(--steel);}}
  .hero-record .breakdown b{{color:var(--chalk-dim); font-weight:500;}}
  @media (max-width:760px){{
    .hero{{grid-template-columns:1fr; text-align:center; padding-top:36px;}}
    .hero-logo{{margin:0 auto;}}
    .hero-record{{border-left:none; border-top:1px solid var(--card-line); padding-left:0; padding-top:20px; text-align:center;}}
  }}

  section{{max-width:1100px; margin:0 auto; padding:52px 24px;}}
  .section-head{{display:flex; align-items:baseline; justify-content:space-between; margin-bottom:26px; gap:16px; flex-wrap:wrap;}}
  .section-head h2{{font-size:28px; margin:0; color:var(--chalk);}}
  .section-head .note{{font-size:13px; color:var(--steel); max-width:360px; text-align:right;}}

  .snapshot{{display:grid; grid-template-columns:repeat(6,1fr); border:1px solid var(--card-line); border-radius:10px; overflow:hidden;}}
  .snap-cell{{padding:20px 14px; text-align:center; border-right:1px solid var(--card-line); background:var(--card);}}
  .snap-cell:last-child{{border-right:none;}}
  .snap-cell .v{{font-size:30px; color:var(--amber);}}
  .snap-cell .k{{font-size:11px; letter-spacing:0.1em; text-transform:uppercase; color:var(--steel); margin-top:6px;}}
  @media (max-width:760px){{
    .snapshot{{grid-template-columns:repeat(3,1fr);}}
    .snap-cell:nth-child(3n){{border-right:none;}}
    .snap-cell:nth-child(3){{border-right:none;}}
  }}

  .history-grid{{display:grid; grid-template-columns:1.3fr 1fr; gap:48px;}}
  .history-copy p{{font-size:15px; line-height:1.75; color:var(--chalk-dim); margin:0 0 16px;}}
  .history-copy p:last-child{{margin-bottom:0;}}
  .timeline{{list-style:none; margin:0; padding:0; border-left:2px solid var(--card-line);}}
  .timeline li{{position:relative; padding:0 0 22px 22px;}}
  .timeline li:last-child{{padding-bottom:0;}}
  .timeline li::before{{content:''; position:absolute; left:-6px; top:4px; width:10px; height:10px; border-radius:50%; background:var(--arena-black); border:2px solid var(--steel);}}
  .timeline li.win::before{{border-color:var(--turf); background:var(--turf);}}
  .timeline li.mark::before{{border-color:var(--amber); background:var(--amber);}}
  .timeline .yr{{font-size:13px; color:var(--amber); letter-spacing:0.06em;}}
  .timeline .ev{{font-size:14px; color:var(--chalk-dim); margin-top:2px; line-height:1.5;}}
  @media (max-width:760px){{.history-grid{{grid-template-columns:1fr;}}}}

  .coaches{{display:flex; flex-wrap:wrap; gap:14px;}}
  .coach-card{{flex:1 1 260px; background:var(--card); border:1px solid var(--card-line); border-radius:10px; padding:18px 20px;}}
  .coach-card .name{{font-size:17px; color:var(--chalk); font-weight:600;}}
  .coach-card .tenure{{font-size:12px; color:var(--steel); margin-top:4px; letter-spacing:0.04em;}}

  .logo-history{{display:flex; flex-wrap:wrap; gap:14px;}}
  .logo-card{{flex:0 0 156px; background:var(--card); border:1px solid var(--card-line); border-radius:10px; padding:20px 16px; text-align:center; display:flex; flex-direction:column; align-items:center;}}
  .logo-card .badge-slot{{height:24px; display:flex; align-items:center; margin-bottom:2px;}}
  .logo-card .badge{{display:inline-block; font-size:10px; letter-spacing:0.08em; text-transform:uppercase; color:var(--amber); background:rgba(232,163,61,0.12); padding:3px 8px; border-radius:4px; font-weight:600;}}
  .logo-card .thumb{{width:88px; height:88px; margin:0 auto 12px; border-radius:50%; background:var(--arena-black-2); border:1px solid var(--card-line); display:flex; align-items:center; justify-content:center; overflow:hidden; flex-shrink:0;}}
  .logo-card .thumb img{{width:100%; height:100%; object-fit:contain; padding:10px;}}
  .logo-card .logo-name{{font-size:14px; color:var(--chalk); font-weight:600; line-height:1.3; min-height:2.6em; display:flex; align-items:flex-start; justify-content:center; margin-bottom:4px;}}
  .logo-card .caption{{font-size:12px; color:var(--steel); line-height:1.4; margin-top:auto;}}

  .records-wrap{{border:1px solid var(--card-line); border-radius:10px; overflow:hidden;}}
  .records-table{{width:100%; border-collapse:collapse;}}
  .records-table thead th{{text-align:left; font-size:11px; letter-spacing:0.1em; text-transform:uppercase; color:var(--steel); padding:12px 18px; border-bottom:1px solid var(--card-line); font-weight:500; background:var(--card);}}
  .records-table thead th:last-child{{text-align:right;}}
  .records-table .group-row td{{padding:10px 18px 8px; font-size:11px; letter-spacing:0.14em; text-transform:uppercase; color:var(--crimson); font-weight:600; background:var(--arena-black-2); border-bottom:1px solid var(--card-line); border-top:1px solid var(--card-line);}}
  .records-table tr.group-row:first-child td{{border-top:none;}}
  .records-table td{{padding:12px 18px; border-bottom:1px solid var(--steel-line); font-size:14px; color:var(--chalk-dim);}}
  .records-table tbody tr:not(.group-row):last-child td{{border-bottom:none;}}
  .records-table td.rec{{color:var(--chalk-dim); padding-left:30px;}}
  .records-table td.holder{{color:var(--amber); font-weight:500;}}
  .records-table td.val{{text-align:right; font-family:'IBM Plex Mono',monospace; color:var(--chalk); white-space:nowrap;}}
  .records-table tbody tr:not(.group-row):hover td{{background:rgba(255,255,255,0.03);}}

  .leaders-grid{{display:grid; grid-template-columns:repeat(3,1fr); gap:14px;}}
  .leader-card{{background:var(--card); border:1px solid var(--card-line); border-radius:10px; padding:20px;}}
  .leader-card .cat{{font-size:11px; letter-spacing:0.12em; text-transform:uppercase; color:var(--crimson); font-weight:600; margin-bottom:12px;}}
  .leader-card .who{{font-size:18px; color:var(--chalk); font-weight:600;}}
  .leader-card .stat{{margin-top:8px; font-size:26px; color:var(--amber); font-family:'IBM Plex Mono',monospace;}}
  .leader-card .stat span{{font-size:13px; color:var(--steel); font-family:'Inter',sans-serif; margin-left:6px;}}
  @media (max-width:760px){{.leaders-grid{{grid-template-columns:1fr 1fr;}}}}
  @media (max-width:480px){{.leaders-grid{{grid-template-columns:1fr;}}}}

  select.unit-select{{
    appearance:none; -webkit-appearance:none; background:var(--card); color:var(--chalk);
    border:1px solid var(--card-line); border-radius:8px; padding:9px 34px 9px 14px;
    font-family:'IBM Plex Mono',monospace; font-size:13px; letter-spacing:0.03em; cursor:pointer;
    background-image:url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='10' height='6'><path d='M0 0L5 6L10 0Z' fill='%2382868c'/></svg>");
    background-repeat:no-repeat; background-position:right 14px center;
  }}
  select.unit-select:focus{{outline:1px solid var(--crimson); outline-offset:1px;}}

  .season-panel{{border:1px solid var(--card-line); border-radius:10px; background:var(--card); padding:26px 28px; display:grid; grid-template-columns:auto 1fr auto; align-items:center; gap:26px;}}
  .season-panel .yr-big{{font-size:44px; color:var(--chalk); line-height:1;}}
  .season-panel .body p{{margin:0; font-size:14px; color:var(--chalk-dim); line-height:1.6; max-width:520px;}}
  .season-panel .rec-block{{text-align:right;}}
  .season-panel .rec-block .rv{{font-size:26px; color:var(--amber); font-family:'IBM Plex Mono',monospace;}}
  .season-panel .rec-block .rk{{font-size:11px; letter-spacing:0.1em; text-transform:uppercase; color:var(--steel); margin-top:4px;}}
  .season-panel .rec-block .po{{font-size:12px; color:var(--turf); margin-top:8px; font-family:'IBM Plex Mono',monospace;}}
  @media (max-width:700px){{
    .season-panel{{grid-template-columns:1fr; text-align:left;}}
    .season-panel .rec-block{{text-align:left;}}
  }}
  .season-games{{margin-top:16px;}}
  .season-games .empty{{padding:22px 24px; text-align:center; font-size:13px; color:var(--steel); border:1px dashed var(--card-line); border-radius:10px;}}

  .schedule{{border:1px solid var(--card-line); border-radius:10px; overflow:hidden;}}
  .game-row{{display:grid; grid-template-columns:96px 1fr 100px 64px; align-items:center; padding:13px 18px; border-bottom:1px solid var(--steel-line); font-size:14px;}}
  .game-row:last-child{{border-bottom:none;}}
  .game-row .date{{color:var(--steel); font-size:12px;}}
  .game-row .opp{{color:var(--chalk-dim);}}
  .game-row .score{{font-family:'IBM Plex Mono',monospace; text-align:right; color:var(--chalk);}}
  .game-row .wl{{text-align:center; font-family:'IBM Plex Mono',monospace; font-size:12px; font-weight:600; padding:3px 0; border-radius:4px; margin-left:16px;}}
  .wl.w{{background:rgba(60,110,71,0.25); color:#7fc492;}}
  .wl.l{{background:rgba(212,38,58,0.2); color:#f0919c;}}
  .game-row.playoff{{background:rgba(232,163,61,0.05);}}
  .game-row.playoff .date{{color:var(--amber);}}

  footer{{max-width:1100px; margin:0 auto; padding:40px 24px 60px; display:flex; justify-content:space-between; flex-wrap:wrap; gap:12px; font-size:12px; color:var(--steel);}}

  ::selection{{background:var(--crimson); color:var(--chalk);}}
</style>
</head>
<body>

<div class="topbar">
  <div class="site"><strong>Indoor Football Index</strong> / Teams</div>
  <nav>
    <a href="leagues.html">Leagues</a>
    <a href="teams.html">Teams</a>
    <a href="index.html#schedule">Schedule</a>
    <a href="index.html#news">News</a>
  </nav>
</div>

<header class="hero">
  <div class="hero-logo">
    <img src="{logo_url}" alt="{team_name} logo">
  </div>
  <div class="hero-id">
    <p class="eyebrow">{league_history}</p>
    <h1 class="display">{team_name_html}</h1>
    <div class="sub">
      <span><b>{arena_current}</b></span>
      {arena_prior_span}
      <span>Est. <b>{est_year}</b></span>
    </div>
  </div>
  <div class="hero-record">
    <p class="label">All-time record</p>
    <div class="num display">{record_overall}</div>
    <div class="breakdown"><b>{record_reg}</b> regular season &middot; <b>{record_post}</b> postseason</div>
  </div>
</header>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1100 14"><line x1="0" y1="7" x2="1100" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

<section id="snapshot-section">
  <div class="snapshot">
    <div class="snap-cell"><div class="v mono">{latest_record}</div><div class="k">{latest_season} record</div></div>
    <div class="snap-cell"><div class="v mono">{championships}</div><div class="k">Championships</div></div>
    <div class="snap-cell"><div class="v mono">{streak}</div><div class="k">Current streak</div></div>
    <div class="snap-cell"><div class="v mono">{playoff_count}</div><div class="k">Playoff berths</div></div>
    <div class="snap-cell"><div class="v mono">{first_playoff_win}</div><div class="k">First playoff win</div></div>
    <div class="snap-cell"><div class="v mono">{seasons_played}</div><div class="k">Seasons played</div></div>
  </div>
</section>

<section id="history">
  <div class="section-head">
    <h2 class="display">Franchise history</h2>
  </div>
  <div class="history-grid">
    <div class="history-copy">
      {history_paragraphs}
    </div>
    <ul class="timeline">
{timeline_items}
    </ul>
  </div>
</section>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1100 14"><line x1="0" y1="7" x2="1100" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

{logo_section}<section id="coaches">
  <div class="section-head"><h2 class="display">Coaching history</h2></div>
  <div class="coaches">
{coaches_html}
  </div>
</section>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1100 14"><line x1="0" y1="7" x2="1100" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

<section id="records">
  <div class="section-head">
    <h2 class="display">Franchise records</h2>
    <p class="note">Franchise record holders.</p>
  </div>
  <div style="margin-bottom:20px;">
    <select class="unit-select" id="records-filter">
      <option value="offense">Offense</option>
      <option value="defense">Defense</option>
      <option value="special">Special teams</option>
    </select>
  </div>
  <div class="records-wrap">
  <table class="records-table">
    <thead><tr><th>Record</th><th>Holder</th><th>Total</th></tr></thead>
{records_table}
  </table>
  </div>
</section>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1100 14"><line x1="0" y1="7" x2="1100" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

<section id="leaders">
  <div class="section-head">
    <h2 class="display">{latest_season} season leaders</h2>
  </div>
  <div style="margin-bottom:20px;">
    <select class="unit-select" id="unit-filter">
      <option value="offense">Offense</option>
      <option value="defense">Defense</option>
      <option value="special">Special teams</option>
    </select>
  </div>
  <div class="leaders-grid" id="leaders-grid">
{leader_cards}
  </div>
</section>

<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1100 14"><line x1="0" y1="7" x2="1100" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>

<section id="schedule">
  <div class="section-head">
    <h2 class="display">Season by season</h2>
    <p class="note">Pick a year for the record and how the season played out.</p>
  </div>
  <div style="margin-bottom:20px;">
    <select class="unit-select" id="season-select">
{year_options}
    </select>
  </div>
  <div class="season-panel">
    <div class="yr-big display" id="season-yr">{default_year}</div>
    <div class="body"><p id="season-note"></p></div>
    <div class="rec-block">
      <div class="rv mono" id="season-rec"></div>
      <div class="rk">Regular season</div>
      <div class="po mono" id="season-post"></div>
    </div>
  </div>
  <div class="season-games" id="season-games"></div>
</section>

<script>
  var seasonData = {{
{season_js}
  }};
  var seasonSelect = document.getElementById('season-select');
  var yrEl = document.getElementById('season-yr');
  var noteEl = document.getElementById('season-note');
  var recEl = document.getElementById('season-rec');
  var postEl = document.getElementById('season-post');
  var gamesEl = document.getElementById('season-games');
  function renderSeason(y){{
    var d = seasonData[y];
    yrEl.textContent = y;
    noteEl.textContent = d.note;
    recEl.innerHTML = d.rec;
    postEl.textContent = d.post;
    postEl.style.display = d.post ? 'block' : 'none';
    if(d.games && d.games.length){{
      var html = '<div class="schedule">';
      d.games.forEach(function(g){{
        var isPlayoff = g[0].indexOf('PO') !== -1;
        var wl = g[3];
        html += '<div class="game-row'+(isPlayoff?' playoff':'')+'">'
          + '<span class="date">'+g[0]+'</span>'
          + '<span class="opp">'+g[1]+'</span>'
          + '<span class="score mono">'+g[2]+'</span>'
          + '<span class="wl '+(wl==='W'?'w':'l')+'">'+wl+'</span>'
          + '</div>';
      }});
      html += '</div>';
      gamesEl.innerHTML = html;
    }} else {{
      gamesEl.innerHTML = '<div class="empty">Game-by-game box scores for '+y+' aren\'t in the archive yet &mdash; season record only.</div>';
    }}
  }}
  seasonSelect.addEventListener('change', function(){{ renderSeason(this.value); }});
  renderSeason('{default_year}');

  var unitSelect = document.getElementById('unit-filter');
  var cards = document.querySelectorAll('#leaders-grid .leader-card');
  unitSelect.addEventListener('change', function(){{
    var val = this.value;
    cards.forEach(function(c){{ c.style.display = (c.getAttribute('data-unit') === val) ? '' : 'none'; }});
  }});

  var recordsSelect = document.getElementById('records-filter');
  var recordBodies = document.querySelectorAll('.records-table tbody');
  recordsSelect.addEventListener('change', function(){{
    var val = this.value;
    recordBodies.forEach(function(tb){{ tb.style.display = (tb.getAttribute('data-unit') === val) ? '' : 'none'; }});
  }});
</script>

<footer>
  <span>Indoor Football Index &middot; Team data compiled from IFL box scores and franchise records.</span>
  <span>{team_name} &middot; IFL</span>
</footer>

</body>
</html>
"""


def generate(xlsx_path, schedule_xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    info = load_team_info(wb)
    team_name = info.get("Team Name") or "Team"

    schedule_wb = openpyxl.load_workbook(schedule_xlsx_path, data_only=True)
    seasons = load_game_history(schedule_wb, team_name)
    season_table = load_season_table(wb, team_name)
    years_sorted = sorted(set(seasons.keys()) | set(season_table.keys() if season_table else []))
    est_year = info.get("Years", "").split("-")[0].strip() if info.get("Years") else (years_sorted[0] if years_sorted else "")

    # hero arena text
    stadium_hist = info.get("Stadium History") or ""
    arenas = [a.strip() for a in stadium_hist.split(",") if a.strip()]
    arena_current = arenas[-1] if arenas else ""
    arena_prior_span = ""
    if len(arenas) > 1:
        arena_prior_span = f'<span>Formerly {esc(", ".join(arenas[:-1]))}</span>'

    playoff_years = [y.strip() for y in (info.get("Playoff appearances") or "").split(",") if y.strip()]

    year_options, default_year = build_year_options(seasons, season_table)
    season_js = build_season_js(seasons, season_table)

    grouped_records = load_records(wb)
    records_table = build_records_table(grouped_records)

    latest_season, leaders, latest_season_raw = load_current_leaders(wb)
    leader_cards = build_leader_cards(leaders)
    latest_season_key = str(int(latest_season_raw)) if latest_season_raw is not None else None

    coaches_html = build_coaches(info.get("Coach History"))

    logo_entries = load_logo_history(info)
    logo_history_html = build_logo_history(logo_entries)
    if logo_entries:
        logo_section = (
            '<section id="logos">\n'
            '  <div class="section-head"><h2 class="display">Logo history</h2></div>\n'
            '  <div class="logo-history">\n'
            f'{logo_history_html}\n'
            '  </div>\n'
            '</section>\n\n'
            '<div class="hashes"><svg preserveAspectRatio="none" viewBox="0 0 1100 14">'
            '<line x1="0" y1="7" x2="1100" y2="7" stroke="#2a2d31" stroke-width="1"/></svg></div>\n\n'
        )
    else:
        logo_section = ""

    history_text = info.get("Team History") or ""
    paragraphs = [p.strip() for p in re.split(r"(?<=[.!?])\s{2,}|\n\n", history_text) if p.strip()]
    if len(paragraphs) <= 1 and history_text:
        # fall back: split on sentence groups of ~3 for readability
        sentences = re.split(r"(?<=[.!?])\s+", history_text)
        paragraphs = []
        chunk = []
        for s in sentences:
            chunk.append(s)
            if len(chunk) == 3:
                paragraphs.append(" ".join(chunk))
                chunk = []
        if chunk:
            paragraphs.append(" ".join(chunk))
    history_paragraphs = "\n      ".join(f"<p>{esc(p)}</p>" for p in paragraphs)

    _, championship_years = parse_championships(info.get("Championships"))
    timeline_items = []
    for y in years_sorted:
        reg, post, _ = resolve_season(y, seasons, season_table)
        won_title = y in championship_years
        cls = ' class="mark"' if won_title else (' class="win"' if post else "")
        note = f"{reg} regular season." + (f" Postseason {post}." if post else "")
        if won_title:
            note += " Won the championship."
        timeline_items.append(
            f'      <li{cls}><span class="yr mono">{esc(y)}</span><div class="ev">{esc(note)}</div></li>'
        )
    timeline_items_html = "\n".join(timeline_items)

    html = TEMPLATE.format(
        team_name=esc(team_name),
        team_name_html=esc(team_name).replace(" ", "<br>", 1) if " " in team_name else esc(team_name),
        logo_url=esc(info.get("Current Logo") or ""),
        league_history=esc(info.get("League History") or ""),
        arena_current=esc(arena_current),
        arena_prior_span=arena_prior_span,
        est_year=esc(est_year),
        record_overall=esc(info.get("Record") or "").replace("-", "&ndash;"),
        record_reg=esc(info.get("Reg. Record") or ""),
        record_post=esc(info.get("Post. Record") or ""),
        latest_season=esc(str(latest_season) if latest_season else ""),
        latest_record=esc(resolve_season(latest_season_key, seasons, season_table)[0]) if latest_season_key else "",
        all_time_record=esc((info.get("Record") or "").replace("-", "\u2013")),
        championships=esc(parse_championships(info.get("Championships"))[0]),
        streak=esc(current_streak(seasons) or "\u2014"),
        playoff_count=len(playoff_years),
        first_playoff_win=esc(first_playoff_win_year(seasons) or "\u2014"),
        seasons_played=len(years_sorted),
        history_paragraphs=history_paragraphs,
        timeline_items=timeline_items_html,
        coaches_html=coaches_html,
        logo_section=logo_section,
        records_table=records_table,
        leader_cards=leader_cards,
        year_options=year_options,
        default_year=esc(default_year),
        season_js=season_js,
    )

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python3 generate_team_page.py <team.xlsx> <schedule.xlsx> <output.html>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2], sys.argv[3])
