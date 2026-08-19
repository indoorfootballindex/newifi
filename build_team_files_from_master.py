#!/usr/bin/env python3
"""
Indoor Football Index — master player stats splitter/merger.

Takes ONE big workbook containing every player's stats — headers Name, Pos,
Team, Season, GP, P Comp, P Att, ... (Team sits in column C, between Pos and
Season) — and routes each row into the right team file. The Team column is
matched by its header text, not its position, so this still works even if
the column order changes later.

  - If that team's .xlsx already exists in the output folder, its Player
    Stats tab gets MERGED — matched on (Name, Season). A matching row gets
    updated in place; a new one gets appended. Every other tab (Team Info,
    Roster, Game History, Franchise Record Holders, Season) is left
    completely untouched.

  - If that team's .xlsx doesn't exist yet, a brand new one is created by
    cloning a template file's structure (so Team Info / Roster / Franchise
    Record Holders start with the correct real headers, not guessed ones)
    and setting its Team Name, with Player Stats populated from the master
    file. The clone starts otherwise blank — you'll still need to fill in
    the rest of Team Info, Roster, etc. by hand.

Usage:
    python3 build_team_files_from_master.py master.xlsx teams_folder template.xlsx
"""

import sys
import os
import re
import shutil
import openpyxl


def slug_filename(team_name):
    safe = re.sub(r"[^A-Za-z0-9]+", "_", str(team_name)).strip("_")
    return safe + ".xlsx"


def load_master_rows(master_path):
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    team_col = None
    for h in idx:
        if str(h).strip().lower() == "team":
            team_col = h
            break
    if team_col is None:
        print("ERROR: no 'Team' column found in the master file's header row.")
        print(f"Headers found: {list(idx.keys())}")
        sys.exit(1)

    by_team = {}
    for r in rows[1:]:
        team = r[idx[team_col]]
        if not team:
            continue
        row_dict = {h: r[i] for h, i in idx.items() if h != team_col}
        if not row_dict.get("Name"):
            continue
        by_team.setdefault(team, []).append(row_dict)
    return by_team


def merge_into_existing(path, new_rows):
    wb = openpyxl.load_workbook(path)
    ws = wb["Player Stats"]
    existing_rows = list(ws.iter_rows(values_only=True))
    headers = list(existing_rows[0])
    header_idx = {h: i for i, h in enumerate(headers) if h is not None}

    keyed = {}
    order = []
    for r in existing_rows[1:]:
        if not r or not r[header_idx.get("Name", 0)]:
            continue
        name = r[header_idx["Name"]]
        season = r[header_idx["Season"]] if "Season" in header_idx else None
        key = (name, season)
        keyed[key] = list(r)
        order.append(key)

    added, updated = 0, 0
    for row_dict in new_rows:
        name = row_dict.get("Name")
        season = row_dict.get("Season")
        key = (name, season)
        row_vals = [None] * len(headers)
        for h, v in row_dict.items():
            if h in header_idx:
                row_vals[header_idx[h]] = v
        if key in keyed:
            keyed[key] = row_vals
            updated += 1
        else:
            keyed[key] = row_vals
            order.append(key)
            added += 1

    ws.delete_rows(2, ws.max_row)
    for key in order:
        ws.append(keyed[key])

    wb.save(path)
    return added, updated


def create_new_team_file(template_path, out_path, team_name, rows):
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)

    # Team Info: keep the header row, blank the rest, set Team Name
    info_ws = wb["Team Info"]
    info_headers = [c.value for c in info_ws[1]]
    info_ws.delete_rows(2, info_ws.max_row)
    blank_row = [None] * len(info_headers)
    if "Team Name" in info_headers:
        blank_row[info_headers.index("Team Name")] = team_name
    info_ws.append(blank_row)

    # Roster / Franchise Record Holders / Game History / Season: keep headers
    # only, clear any template data rows so nothing leaks from the template.
    for tab in ("Roster", "Franchise Record Holders", "Game History", "Season"):
        if tab in wb.sheetnames:
            ws = wb[tab]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)

    # Player Stats: keep headers, replace all data with this team's rows
    ps_ws = wb["Player Stats"]
    ps_headers = [c.value for c in ps_ws[1]]
    ps_ws.delete_rows(2, ps_ws.max_row)
    for row_dict in rows:
        row_vals = [row_dict.get(h) for h in ps_headers]
        ps_ws.append(row_vals)

    wb.save(out_path)


def main():
    if len(sys.argv) != 4:
        print("Usage: python3 build_team_files_from_master.py <master.xlsx> <teams_folder> <template.xlsx>")
        sys.exit(1)
    master_path, teams_folder, template_path = sys.argv[1], sys.argv[2], sys.argv[3]

    by_team = load_master_rows(master_path)
    os.makedirs(teams_folder, exist_ok=True)

    print(f"Found {len(by_team)} teams in the master file:\n")
    for team, rows in sorted(by_team.items()):
        out_path = os.path.join(teams_folder, slug_filename(team))
        if os.path.exists(out_path):
            added, updated = merge_into_existing(out_path, rows)
            print(f"  {team}: updated existing file ({added} new, {updated} updated player-seasons)")
        else:
            create_new_team_file(template_path, out_path, team, rows)
            print(f"  {team}: created new file ({len(rows)} player-seasons) — fill in Team Info/Roster/etc. by hand")


if __name__ == "__main__":
    main()
